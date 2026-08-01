import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

from database.db import get_connection


# =========================================================
# CATÁLOGOS
# =========================================================
ESTADO_PAGO_ANULADO = 0
ESTADO_PAGO_REGISTRADO = 1

ESTADO_OPERACION_SERVICIO_CANCELADO = 4

TIPO_OPERACION_NORMAL = 1
TIPO_OPERACION_CONTRATO = 2

CLASE_CONTRATO_ESTANDAR = 1
CLASE_CONTRATO_ESPECIAL = 2

METODOS_PAGO = {
    1: "Efectivo",
    2: "QR",
}


# =========================================================
# HELPERS
# =========================================================
def nombre_metodo_pago(metodo_pago):
    try:
        return METODOS_PAGO.get(int(metodo_pago), str(metodo_pago))
    except Exception:
        return str(metodo_pago or "")


def es_id_valido(valor):
    return valor not in (None, "", 0, "0")


def convertir_float(valor):
    try:
        return float(valor or 0)
    except Exception:
        return 0.0


def columna_existe(cursor, tabla, columna):
    try:
        cursor.execute(f"PRAGMA table_info({tabla})")
        return columna in [row["name"] for row in cursor.fetchall()]
    except Exception:
        return False


def tabla_existe(cursor, tabla):
    try:
        cursor.execute(
            "SELECT COUNT(*) AS Total FROM sqlite_master WHERE type='table' AND name = ?",
            (tabla,),
        )
        row = cursor.fetchone()
        return int(row["Total"] if row else 0) > 0
    except Exception:
        return False


def placas_contrato(cursor, contrato_id, placa_fallback="-"):
    """
    Devuelve todas las placas asociadas a un contrato.
    Usa CONTRATOVEHICULO si existe; si no hay datos, usa la placa principal.
    """
    if not es_id_valido(contrato_id):
        return placa_fallback or "-"

    placas = []

    if tabla_existe(cursor, "CONTRATOVEHICULO"):
        cursor.execute("""
            SELECT V.Placa
            FROM CONTRATOVEHICULO CV
            INNER JOIN VEHICULO V ON V.Vehiculo = CV.Vehiculo
            WHERE CV.Contrato = ?
              AND CV.Estado = 1
            ORDER BY V.Placa ASC
        """, (contrato_id,))
        placas = [str(row["Placa"] or "").strip() for row in cursor.fetchall() if str(row["Placa"] or "").strip()]

    if placas:
        return ", ".join(placas)

    return placa_fallback or "-"


def formatear_fecha(valor):
    """
    Muestra fechas siempre como DD/MM/YYYY o DD/MM/YYYY HH:MM.
    Acepta fechas desde SQLite en YYYY-MM-DD, YYYY-MM-DD HH:MM:SS,
    o fechas que ya vengan como DD/MM/YYYY.
    """
    if not valor:
        return ""

    texto = str(valor).strip()

    formatos = [
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y-%m-%d",
        "%d/%m/%Y %H:%M:%S",
        "%d/%m/%Y %H:%M",
        "%d/%m/%Y",
    ]

    for formato in formatos:
        try:
            if "%H" in formato:
                fecha = datetime.strptime(texto[:19], formato)
                return fecha.strftime("%d/%m/%Y %H:%M")

            fecha = datetime.strptime(texto[:10], formato)
            return fecha.strftime("%d/%m/%Y")
        except Exception:
            pass

    if "-" in texto and len(texto) >= 10:
        partes = texto[:10].split("-")
        if len(partes) == 3:
            resto = texto[10:16]
            return f"{partes[2]}/{partes[1]}/{partes[0]}{resto}"

    return texto


def convertir_fecha_busqueda_a_bd(valor):
    """
    Convierte DD/MM/YYYY a YYYY-MM-DD para buscar correctamente en SQLite.
    """
    valor = str(valor or "").strip()

    if not valor:
        return ""

    try:
        return datetime.strptime(valor, "%d/%m/%Y").strftime("%Y-%m-%d")
    except ValueError:
        return None


def convertir_fecha_a_datetime(valor):
    if not valor:
        return None

    texto = str(valor).strip()

    formatos = [
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y-%m-%d",
        "%d/%m/%Y %H:%M:%S",
        "%d/%m/%Y %H:%M",
        "%d/%m/%Y",
    ]

    for formato in formatos:
        try:
            return datetime.strptime(texto[:19], formato)
        except Exception:
            pass

    return None


def calcular_meses_entre_fechas(fecha_inicio, fecha_fin):
    """
    Calcula meses para contratos.

    Ejemplo:
    Inicio: 09/05/2026
    Fin:    08/06/2026
    Resultado: 1 mes
    """
    inicio = convertir_fecha_a_datetime(fecha_inicio)
    fin = convertir_fecha_a_datetime(fecha_fin)

    if not inicio or not fin:
        return 1

    meses = (fin.year - inicio.year) * 12 + (fin.month - inicio.month)

    # En contratos, la fecha fin normalmente es un día antes de cumplirse el mes.
    # Si termina el mismo día o después, cuenta como un mes adicional.
    # Si termina justo un día antes, el cálculo mensual ya es correcto.
    if fin.day >= inicio.day:
        meses += 1

    if meses <= 0:
        meses = 1

    return meses


def texto_meses(cantidad):
    try:
        cantidad = int(cantidad or 1)
    except Exception:
        cantidad = 1

    if cantidad <= 1:
        return "1 mes"
    return f"{cantidad} meses"


def limpiar_placa(valor):
    return str(valor or "").replace(" ", "").replace("-", "").upper().strip()


def configurar_treeview():
    style = ttk.Style()
    try:
        style.theme_use("default")
    except Exception:
        pass

    style.configure(
        "Reportes.Treeview",
        background="white",
        foreground="#111827",
        rowheight=28,
        fieldbackground="white",
        borderwidth=0,
        relief="flat",
        font=("Arial", 10),
    )

    style.configure(
        "Reportes.Treeview.Heading",
        background="#f3f4f6",
        foreground="#111827",
        font=("Arial", 10, "bold"),
        borderwidth=0,
        relief="flat",
        padding=(5, 8),
    )

    style.map(
        "Reportes.Treeview",
        background=[("selected", "#e5e7eb")],
        foreground=[("selected", "#111827")],
    )


class ReportsView:
    def __init__(self, parent):
        self.parent = parent

        self.entry_from = None
        self.entry_to = None
        self.entry_plate = None
        self.tree = None
        self.total_label = None

        self.current_rows = []

    def build(self):
        configurar_treeview()
        self.build_filters()
        self.build_table()
        self.build_footer()
        self.load_reports()

    # =====================================================
    # INTERFAZ
    # =====================================================
    def build_filters(self):
        filters_frame = tk.Frame(self.parent, bg="white")
        filters_frame.pack(fill="x", padx=15, pady=15)

        tk.Label(
            filters_frame,
            text="Desde (DD/MM/YYYY):",
            font=("Arial", 10, "bold"),
            bg="white",
            fg="#111827"
        ).grid(row=0, column=0, padx=(5, 5), pady=5, sticky="w")

        self.entry_from = tk.Entry(filters_frame, font=("Arial", 10), width=14)
        self.entry_from.grid(row=0, column=1, padx=(0, 12), pady=5, sticky="w")

        tk.Label(
            filters_frame,
            text="Hasta (DD/MM/YYYY):",
            font=("Arial", 10, "bold"),
            bg="white",
            fg="#111827"
        ).grid(row=0, column=2, padx=(5, 5), pady=5, sticky="w")

        self.entry_to = tk.Entry(filters_frame, font=("Arial", 10), width=14)
        self.entry_to.grid(row=0, column=3, padx=(0, 12), pady=5, sticky="w")

        tk.Label(
            filters_frame,
            text="Placa:",
            font=("Arial", 10, "bold"),
            bg="white",
            fg="#111827"
        ).grid(row=0, column=4, padx=(5, 5), pady=5, sticky="w")

        self.entry_plate = tk.Entry(filters_frame, font=("Arial", 10), width=14)
        self.entry_plate.grid(row=0, column=5, padx=(0, 12), pady=5, sticky="w")
        self.entry_plate.bind("<KeyRelease>", lambda event: self.load_reports())

        search_button = tk.Button(
            filters_frame,
            text="Buscar",
            font=("Arial", 10, "bold"),
            bg="#111827",
            fg="white",
            activebackground="#374151",
            activeforeground="white",
            bd=0,
            relief="flat",
            padx=14,
            pady=6,
            cursor="hand2",
            command=self.load_reports
        )
        search_button.grid(row=0, column=6, padx=8, pady=5)

        clear_button = tk.Button(
            filters_frame,
            text="Limpiar",
            font=("Arial", 10, "bold"),
            bg="#6b7280",
            fg="white",
            activebackground="#4b5563",
            activeforeground="white",
            bd=0,
            relief="flat",
            padx=14,
            pady=6,
            cursor="hand2",
            command=self.clear_filters
        )
        clear_button.grid(row=0, column=7, padx=8, pady=5)

        export_button = tk.Button(
            filters_frame,
            text="Exportar Excel",
            font=("Arial", 10, "bold"),
            bg="#16a34a",
            fg="white",
            activebackground="#15803d",
            activeforeground="white",
            bd=0,
            relief="flat",
            padx=14,
            pady=6,
            cursor="hand2",
            command=self.export_excel
        )
        export_button.grid(row=0, column=8, padx=8, pady=5)

    def build_table(self):
        table_frame = tk.Frame(self.parent, bg="white")
        table_frame.pack(fill="both", expand=True, padx=15, pady=(0, 10))

        columns = (
            "FechaPago",
            "Placa(s)",
            "Detalle",
            "Tiempo",
            "MetodoPago",
            "Empleado",
            "Total"
        )

        self.tree = ttk.Treeview(
            table_frame,
            columns=columns,
            show="headings",
            height=18,
            style="Reportes.Treeview"
        )

        self.tree.heading("FechaPago", text="Fecha pago")
        self.tree.heading("Placa(s)", text="Placa(s)")
        self.tree.heading("Detalle", text="Detalle")
        self.tree.heading("Tiempo", text="Tiempo")
        self.tree.heading("MetodoPago", text="Método")
        self.tree.heading("Empleado", text="Empleado")
        self.tree.heading("Total", text="Total")

        self.tree.column("FechaPago", width=150, anchor="center", stretch=False)
        self.tree.column("Placa(s)", width=170, anchor="center", stretch=False)
        self.tree.column("Detalle", width=360, anchor="w", stretch=True)
        self.tree.column("Tiempo", width=120, anchor="center", stretch=False)
        self.tree.column("MetodoPago", width=120, anchor="center", stretch=False)
        self.tree.column("Empleado", width=180, anchor="w", stretch=True)
        self.tree.column("Total", width=120, anchor="e", stretch=False)

        scrollbar_y = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)

        self.tree.configure(yscrollcommand=scrollbar_y.set)

        self.tree.grid(row=0, column=0, sticky="nsew")
        scrollbar_y.grid(row=0, column=1, sticky="ns")

        table_frame.grid_rowconfigure(0, weight=1)
        table_frame.grid_columnconfigure(0, weight=1)

    def build_footer(self):
        footer = tk.Frame(self.parent, bg="white")
        footer.pack(fill="x", padx=15, pady=(0, 15))

        self.total_label = tk.Label(
            footer,
            text="Total: Bs 0.00",
            font=("Arial", 12, "bold"),
            bg="white",
            fg="#111827"
        )
        self.total_label.pack(side="right")

    # =====================================================
    # FILTROS
    # =====================================================
    def clear_filters(self):
        self.entry_from.delete(0, tk.END)
        self.entry_to.delete(0, tk.END)
        self.entry_plate.delete(0, tk.END)
        self.load_reports()

    def validar_rango_fechas(self, date_from_text, date_to_text):
        date_from_bd = convertir_fecha_busqueda_a_bd(date_from_text)
        date_to_bd = convertir_fecha_busqueda_a_bd(date_to_text)

        if date_from_text and date_from_bd is None:
            messagebox.showwarning(
                "Fecha inválida",
                "La fecha 'Desde' debe estar en formato DD/MM/YYYY.\nEjemplo: 29/04/2026"
            )
            return None, None, False

        if date_to_text and date_to_bd is None:
            messagebox.showwarning(
                "Fecha inválida",
                "La fecha 'Hasta' debe estar en formato DD/MM/YYYY.\nEjemplo: 29/04/2026"
            )
            return None, None, False

        if date_from_bd and date_to_bd and date_from_bd > date_to_bd:
            messagebox.showwarning(
                "Rango inválido",
                "La fecha 'Desde' no puede ser mayor que la fecha 'Hasta'."
            )
            return None, None, False

        return date_from_bd, date_to_bd, True

    def format_duration(self, minutes):
        minutes = int(minutes or 0)
        horas = minutes // 60
        mins = minutes % 60

        if horas > 0 and mins > 0:
            return f"{horas} h {mins} min"
        if horas > 0:
            return f"{horas} h"
        return f"{mins} min"

    # =====================================================
    # DETALLES
    # =====================================================
    def get_operation_services(self, cursor, operation_id):
        if not es_id_valido(operation_id):
            return "Parqueo"

        cursor.execute("""
            SELECT S.Nombre
            FROM OPERACIONSERVICIO OS
            INNER JOIN SERVICIO S ON OS.Servicio = S.Servicio
            WHERE OS.Operacion = ?
              AND OS.Estado != ?
            ORDER BY S.Nombre ASC
        """, (operation_id, ESTADO_OPERACION_SERVICIO_CANCELADO))

        rows = cursor.fetchall()
        names = [r["Nombre"] for r in rows]

        if not names:
            return "Parqueo"

        return "Parqueo, " + ", ".join(names)

    def get_operation_services_only(self, cursor, operation_id):
        if not es_id_valido(operation_id):
            return ""

        cursor.execute("""
            SELECT S.Nombre
            FROM OPERACIONSERVICIO OS
            INNER JOIN SERVICIO S ON OS.Servicio = S.Servicio
            WHERE OS.Operacion = ?
              AND OS.Estado != ?
            ORDER BY S.Nombre ASC
        """, (operation_id, ESTADO_OPERACION_SERVICIO_CANCELADO))

        rows = cursor.fetchall()
        return ", ".join([r["Nombre"] for r in rows])

    def detalle_contrato(self, row):
        codigo = row["CodigoContrato"] or (
            f"Contrato #{row['ContratoReal']}" if es_id_valido(row["ContratoReal"]) else "Contrato"
        )

        clase = int(row["ClaseContrato"] or 0)
        horas = row["HorasPermitidasDia"]

        if clase == CLASE_CONTRATO_ESPECIAL:
            return f"Contrato especial {codigo}"

        if horas:
            try:
                return f"Contrato {int(horas)}h {codigo}"
            except Exception:
                return f"Contrato {horas}h {codigo}"

        return f"Contrato {codigo}"

    def tiempo_contrato(self, row):
        duracion_mes = row["DuracionMes"]
        if duracion_mes not in (None, "", 0, "0"):
            return texto_meses(duracion_mes)

        meses = calcular_meses_entre_fechas(
            row["ContratoFechaInicio"],
            row["ContratoFechaFin"]
        )
        return texto_meses(meses)

    # =====================================================
    # CARGA DE DATOS
    # =====================================================
    def load_reports(self):
        if not self.tree:
            return

        for item in self.tree.get_children():
            self.tree.delete(item)

        date_from_text = self.entry_from.get().strip()
        date_to_text = self.entry_to.get().strip()
        plate = self.entry_plate.get().strip().upper()

        date_from, date_to, ok = self.validar_rango_fechas(date_from_text, date_to_text)
        if not ok:
            return

        conn = get_connection()
        cursor = conn.cursor()

        try:
            pago_tiene_contrato = columna_existe(cursor, "PAGO", "Contrato")
            existe_contrato_vehiculo = tabla_existe(cursor, "CONTRATOVEHICULO")

            pago_contrato_select = "P.Contrato AS PagoContrato," if pago_tiene_contrato else "NULL AS PagoContrato,"
            pago_contrato_join = "CD.Contrato = P.Contrato" if pago_tiene_contrato else "1 = 0"

            query = f"""
                SELECT
                    P.Pago,
                    P.Operacion AS PagoOperacion,
                    {pago_contrato_select}
                    P.FechaPago,
                    P.MetodoPago,
                    P.Monto AS MontoPago,
                    P.Estado AS EstadoPago,

                    O.Operacion AS OperacionReal,
                    O.CodigoOperacion,
                    O.Contrato AS OperacionContrato,
                    O.TipoOperacion,
                    O.MinutosEstadia,
                    O.FechaIngreso,
                    O.FechaSalida,
                    O.MontoParqueo,
                    O.MontoServicios,
                    O.MontoTotal,

                    COALESCE(CO.Contrato, CD.Contrato) AS ContratoReal,
                    COALESCE(CO.CodigoContrato, CD.CodigoContrato) AS CodigoContrato,
                    COALESCE(CO.MontoContrato, CD.MontoContrato) AS MontoContrato,
                    COALESCE(CO.DuracionMes, CD.DuracionMes) AS DuracionMes,
                    COALESCE(CO.ClaseContrato, CD.ClaseContrato) AS ClaseContrato,
                    COALESCE(CO.HorasPermitidasDia, CD.HorasPermitidasDia) AS HorasPermitidasDia,

                    COALESCE(CO.FechaInicio, CD.FechaInicio) AS ContratoFechaInicio,
                    COALESCE(CO.FechaFin, CD.FechaFin) AS ContratoFechaFin,

                    V.Placa,

                    U.Nombre AS Empleado
                FROM PAGO P
                LEFT JOIN OPERACION O
                       ON O.Operacion = P.Operacion

                LEFT JOIN CONTRATO CO
                       ON CO.Contrato = O.Contrato

                LEFT JOIN CONTRATO CD
                       ON {pago_contrato_join}
                      AND O.Operacion IS NULL

                LEFT JOIN VEHICULO V
                       ON V.Vehiculo = COALESCE(O.Vehiculo, CO.Vehiculo, CD.Vehiculo)

                LEFT JOIN USUARIO U
                       ON U.Usuario = COALESCE(O.UsuarioSalida, O.UsuarioIngreso, P.Usuario)

                WHERE P.Estado = ?
            """

            params = [ESTADO_PAGO_REGISTRADO]

            if date_from:
                query += " AND date(P.FechaPago) >= date(?)"
                params.append(date_from)

            if date_to:
                query += " AND date(P.FechaPago) <= date(?)"
                params.append(date_to)

            if plate:
                placa_limpia = limpiar_placa(plate)

                if existe_contrato_vehiculo:
                    query += """
                        AND (
                            REPLACE(REPLACE(UPPER(COALESCE(V.Placa, '')), ' ', ''), '-', '') LIKE ?
                            OR EXISTS (
                                SELECT 1
                                FROM CONTRATOVEHICULO CVF
                                INNER JOIN VEHICULO VF ON VF.Vehiculo = CVF.Vehiculo
                                WHERE CVF.Contrato = COALESCE(CO.Contrato, CD.Contrato)
                                  AND CVF.Estado = 1
                                  AND REPLACE(REPLACE(UPPER(COALESCE(VF.Placa, '')), ' ', ''), '-', '') LIKE ?
                            )
                        )
                    """
                    params.extend([f"%{placa_limpia}%", f"%{placa_limpia}%"])
                else:
                    query += """
                        AND REPLACE(REPLACE(UPPER(COALESCE(V.Placa, '')), ' ', ''), '-', '') LIKE ?
                    """
                    params.append(f"%{placa_limpia}%")

            query += " ORDER BY P.FechaPago DESC, P.Pago DESC"

            cursor.execute(query, params)
            rows = cursor.fetchall()

            self.current_rows = []
            total_general = 0.0

            for row in rows:
                operation_id = row["OperacionReal"]
                contrato_id = row["ContratoReal"]

                empleado = row["Empleado"] if row["Empleado"] else "-"

                tipo_operacion_bd = int(row["TipoOperacion"] or TIPO_OPERACION_NORMAL)
                pago_contrato_directo = es_id_valido(row["PagoContrato"]) if pago_tiene_contrato else False
                operacion_con_contrato = (
                    es_id_valido(operation_id)
                    and tipo_operacion_bd == TIPO_OPERACION_CONTRATO
                    and es_id_valido(contrato_id)
                )

                metodo_pago = nombre_metodo_pago(row["MetodoPago"])
                fecha_pago = formatear_fecha(row["FechaPago"])

                monto_pago = convertir_float(row["MontoPago"])
                monto_parqueo_bd = convertir_float(row["MontoParqueo"])
                monto_servicios = convertir_float(row["MontoServicios"])
                monto_total_bd = convertir_float(row["MontoTotal"])

                if pago_contrato_directo:
                    # Pago creado directamente desde Contratos / payment.py.
                    placa = placas_contrato(cursor, contrato_id, row["Placa"] or "-")
                    detalle = self.detalle_contrato(row)
                    tiempo = self.tiempo_contrato(row)
                    total = monto_pago

                elif operacion_con_contrato:
                    # Operación de servicios asociada a un contrato.
                    # Aquí se muestra la placa específica que ingresó al garaje,
                    # no necesariamente todas las placas del contrato.
                    placa = row["Placa"] or "-"

                    servicios = self.get_operation_services_only(cursor, operation_id)
                    codigo = row["CodigoContrato"] or ""
                    if servicios:
                        detalle = f"Servicios con contrato {codigo}: {servicios}".strip()
                    else:
                        detalle = f"Servicios con contrato {codigo}".strip()

                    minutos = int(row["MinutosEstadia"] or 0)
                    tiempo = self.format_duration(minutos)
                    total = monto_pago if monto_pago > 0 else monto_servicios

                else:
                    # Parqueo normal.
                    placa = row["Placa"] or "-"
                    servicios = self.get_operation_services(cursor, operation_id)
                    detalle = servicios

                    minutos = int(row["MinutosEstadia"] or 0)
                    tiempo = self.format_duration(minutos)

                    if monto_parqueo_bd <= 0 and monto_servicios <= 0:
                        total = monto_total_bd if monto_total_bd > 0 else monto_pago
                    else:
                        total = monto_parqueo_bd + monto_servicios

                view_row = (
                    fecha_pago,
                    placa,
                    detalle,
                    tiempo,
                    metodo_pago,
                    empleado,
                    f"Bs {total:.2f}"
                )

                self.current_rows.append({
                    "fecha_pago": fecha_pago,
                    "placa": placa,
                    "detalle": detalle,
                    "tiempo": tiempo,
                    "metodo_pago": metodo_pago,
                    "empleado": empleado,
                    "total": total
                })

                self.tree.insert("", "end", values=view_row)
                total_general += total

            self.total_label.config(text=f"Total: Bs {total_general:.2f}")

        except Exception as e:
            messagebox.showerror("Error", f"No se pudieron cargar los reportes.\n{str(e)}")
        finally:
            conn.close()

    # =====================================================
    # EXPORTAR EXCEL
    # =====================================================
    def export_excel(self):
        if not self.current_rows:
            messagebox.showwarning("Sin datos", "No hay datos para exportar.")
            return

        file_path = filedialog.asksaveasfilename(
            title="Guardar reporte",
            defaultextension=".xlsx",
            filetypes=[("Archivos Excel", "*.xlsx")],
            initialfile=f"reporte_garaje_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )

        if not file_path:
            return

        try:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Reportes"

            headers = [
                "Fecha pago",
                "Placa(s)",
                "Detalle",
                "Tiempo",
                "Método",
                "Empleado",
                "Total"
            ]
            sheet.append(headers)

            for col in range(1, len(headers) + 1):
                cell = sheet.cell(row=1, column=col)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")

            for row in self.current_rows:
                sheet.append([
                    row["fecha_pago"],
                    row["placa"],
                    row["detalle"],
                    row["tiempo"],
                    row["metodo_pago"],
                    row["empleado"],
                    row["total"]
                ])

            total_general = sum(row["total"] for row in self.current_rows)

            last_row = sheet.max_row + 2
            sheet.cell(row=last_row, column=6, value="Total")
            sheet.cell(row=last_row, column=6).font = Font(bold=True)
            sheet.cell(row=last_row, column=6).alignment = Alignment(horizontal="right")

            sheet.cell(row=last_row, column=7, value=total_general)
            sheet.cell(row=last_row, column=7).font = Font(bold=True)

            widths = {
                "A": 22,
                "B": 24,
                "C": 44,
                "D": 16,
                "E": 16,
                "F": 24,
                "G": 16
            }

            for col_letter, width in widths.items():
                sheet.column_dimensions[col_letter].width = width

            workbook.save(file_path)

            messagebox.showinfo(
                "Exportación exitosa",
                "El reporte se exportó correctamente a Excel."
            )

        except Exception as e:
            messagebox.showerror("Error", f"No se pudo exportar el archivo.\n{str(e)}")
